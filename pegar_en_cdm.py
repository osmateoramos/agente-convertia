"""
PRUEBA: pega el brutas y gestionventas de wowperu ya limpios (ver
limpiar_reportes_convertia.py) en el archivo maestro
'08_CDM_UNIVERSAL_WOW PERU CPA Agosto 23.xlsx', reemplazando el historico
acumulado por los datos nuevos y desplegando las formulas de la fila 2 hasta
donde llegue la data nueva.

Edita SOLO el XML interno de las 3 hojas involucradas (sheet19.xml,
sheet20.xml, sheet21.xml dentro del .xlsx, que es un zip) y copia el resto
del archivo (las otras 35 hojas, estilos, tablas dinamicas, etc.) byte por
byte sin tocarlo. Cargar el libro entero con openpyxl tardaba mas de 19
minutos sin terminar de cargar siquiera -- este enfoque es de segundos.

Mapeo de columnas confirmado a mano (25/08/2026) comparando encabezado por
encabezado entre Procesados/BRUTAS_wowperu y la hoja 'TABLA DE RESULTADOS
INCONCERT': las columnas A:AN (40) coinciden; el reporte actual de Convertia
ya NO trae las columnas AO:BG (19: METODO_PAGO, FOLIO, datos de oferta/addon,
donativos TDC, Partner, cedula, etc.) -- quedan en blanco a proposito.

Uso:
    python pegar_en_cdm.py
"""

import zipfile
import shutil
from pathlib import Path

from lxml import etree
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.datetime import to_excel
import openpyxl

CARPETA = Path(__file__).resolve().parent
ARCHIVO_CDM = CARPETA / "Libros de Cuenta" / "08_CDM_UNIVERSAL_WOW PERU CPA Agosto 23.xlsx"
BRUTAS = CARPETA / "Procesados" / "BRUTAS_wowperu_2026-08-25.xlsx"
GESTION = CARPETA / "Procesados" / "GESTIONVENTAS_wowperu_2026-08-25.xlsx"

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NSMAP = {"m": MAIN_NS}


def q(tag: str) -> str:
    return f"{{{MAIN_NS}}}{tag}"


HOJAS = {
    "TABLA DE RESULTADOS INCONCERT": "xl/worksheets/sheet19.xml",
    "INCONCERT MULTIVENTAS": "xl/worksheets/sheet20.xml",
    "Detalle": "xl/worksheets/sheet21.xml",
}

COLUMNAS_BRUTAS_QUE_COINCIDEN = 40  # A:AN en 'TABLA DE RESULTADOS INCONCERT'
COLUMNAS_BRUTAS_TOTALES_HOJA = 81   # A:CC (AO:CC queda en blanco, no existe en el reporte actual)
COLUMNAS_DETALLE = 24               # A:X

LIMITE_BORRADO = {
    "TABLA DE RESULTADOS INCONCERT": 21869,
    "INCONCERT MULTIVENTAS": 4057,
    "Detalle": 4057,
}


def leer_datos(ruta: Path) -> list[tuple]:
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return filas


def indexar_filas(sheet_data) -> dict[int, object]:
    indice = {}
    for row in sheet_data:
        r = row.get("r")
        if r is not None:
            indice[int(r)] = row
    return indice


def get_row(indice_filas: dict, fila: int):
    row = indice_filas.get(fila)
    if row is None:
        raise ValueError(f"No existe la fila {fila} en el XML -- se esperaba que ya existiera (rango cubierto por datos viejos).")
    return row


def indexar_celdas(row_elem) -> dict[int, object]:
    indice = {}
    for c in row_elem:
        r_attr = c.get("r")
        if r_attr:
            indice[column_index_from_string("".join(ch for ch in r_attr if ch.isalpha()))] = c
    return indice


def get_or_create_cell(row_elem, indice_celdas: dict, col: int, fila: int):
    celda = indice_celdas.get(col)
    if celda is not None:
        return celda
    letra = get_column_letter(col)
    ref = f"{letra}{fila}"
    nueva = etree.Element(q("c"))
    nueva.set("r", ref)
    insertado = False
    for hijo in row_elem:
        r_attr = hijo.get("r")
        if r_attr:
            col_existente = column_index_from_string("".join(ch for ch in r_attr if ch.isalpha()))
            if col_existente > col:
                hijo.addprevious(nueva)
                insertado = True
                break
    if not insertado:
        row_elem.append(nueva)
    indice_celdas[col] = nueva
    return nueva


def limpiar_celda(celda) -> None:
    for hijo in list(celda):
        celda.remove(hijo)
    if "t" in celda.attrib:
        del celda.attrib["t"]


def escribir_valor(celda, valor) -> None:
    limpiar_celda(celda)
    if valor is None or valor == "":
        return
    if hasattr(valor, "date") and hasattr(valor, "hour"):  # datetime
        v = etree.SubElement(celda, q("v"))
        v.text = repr(to_excel(valor))
    elif hasattr(valor, "isoformat") and not hasattr(valor, "hour"):  # date
        v = etree.SubElement(celda, q("v"))
        v.text = repr(to_excel(valor))
    elif isinstance(valor, (int, float)):
        v = etree.SubElement(celda, q("v"))
        v.text = repr(valor)
    else:
        celda.set("t", "inlineStr")
        is_el = etree.SubElement(celda, q("is"))
        t_el = etree.SubElement(is_el, q("t"))
        texto = str(valor)
        if texto != texto.strip():
            t_el.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        t_el.text = texto


def escribir_formula(celda, formula_con_igual: str) -> None:
    limpiar_celda(celda)
    f_el = etree.SubElement(celda, q("f"))
    f_el.text = formula_con_igual[1:] if formula_con_igual.startswith("=") else formula_con_igual


def trasladar_formula(formula: str, fila_origen: int, fila_destino: int) -> str:
    from openpyxl.formula.translate import Translator
    # Translator necesita una celda de origen cualquiera; usamos A como referencia de columna dummy,
    # el metodo solo traslada filas/columnas segun el offset entre origen y destino.
    return Translator(f"={formula}" if not formula.startswith("=") else formula,
                       origin=f"A{fila_origen}").translate_formula(f"A{fila_destino}")


def limpiar_fila_sobrante(indice_filas: dict, fila_num: int, max_col: int) -> None:
    row_elem = indice_filas.get(fila_num)
    if row_elem is None:
        return
    for c in list(row_elem):
        limpiar_celda(c)


def procesar_hoja(sheet_root, nombre_hoja: str, datos, es_gestion: bool) -> int:
    sheet_data = sheet_root.find(q("sheetData"))
    limite_viejo = LIMITE_BORRADO[nombre_hoja]
    indice_filas = indexar_filas(sheet_data)

    if nombre_hoja == "Detalle":
        # pegado directo, sin formulas, columnas A:X
        for i, registro in enumerate(datos):
            fila_num = 2 + i
            row_elem = get_row(indice_filas, fila_num)
            indice_celdas = indexar_celdas(row_elem)
            for col in range(1, COLUMNAS_DETALLE + 1):
                celda = get_or_create_cell(row_elem, indice_celdas, col, fila_num)
                valor = registro[col - 1] if col - 1 < len(registro) else None
                escribir_valor(celda, valor)
        fila_final = 2 + len(datos) - 1
        for fila_num in range(fila_final + 1, limite_viejo + 1):
            limpiar_fila_sobrante(indice_filas, fila_num, COLUMNAS_DETALLE)
        return fila_final

    # TABLA DE RESULTADOS INCONCERT o INCONCERT MULTIVENTAS: fila 2 conserva su formula maestra
    fila2 = get_row(indice_filas, 2)
    indice_celdas_fila2 = indexar_celdas(fila2)
    max_col_fila = max(indice_celdas_fila2.keys())

    if nombre_hoja == "TABLA DE RESULTADOS INCONCERT":
        col_formula_desde = COLUMNAS_BRUTAS_TOTALES_HOJA + 1  # CD
    else:
        col_formula_desde = 1  # toda la fila de INCONCERT MULTIVENTAS es formula

    formulas_maestras = {}
    for col in range(col_formula_desde, max_col_fila + 1):
        celda = indice_celdas_fila2.get(col)
        if celda is None:
            continue
        f_el = celda.find(q("f"))
        if f_el is not None and f_el.text:
            formulas_maestras[col] = f_el.text

    if nombre_hoja == "TABLA DE RESULTADOS INCONCERT":
        # fila 2: pisar los datos crudos A:AN con el primer registro nuevo; AO:CC en blanco
        registro0 = datos[0] if datos else None
        for col in range(1, COLUMNAS_BRUTAS_TOTALES_HOJA + 1):
            celda = get_or_create_cell(fila2, indice_celdas_fila2, col, 2)
            if col <= COLUMNAS_BRUTAS_QUE_COINCIDEN and registro0 is not None:
                escribir_valor(celda, registro0[col - 1] if col - 1 < len(registro0) else None)
            else:
                limpiar_celda(celda)

        for i in range(1, len(datos)):
            fila_num = 2 + i
            row_elem = get_row(indice_filas, fila_num)
            indice_celdas = indexar_celdas(row_elem)
            registro = datos[i]
            for col in range(1, COLUMNAS_BRUTAS_TOTALES_HOJA + 1):
                celda = get_or_create_cell(row_elem, indice_celdas, col, fila_num)
                if col <= COLUMNAS_BRUTAS_QUE_COINCIDEN:
                    escribir_valor(celda, registro[col - 1] if col - 1 < len(registro) else None)
                else:
                    limpiar_celda(celda)
            for col, formula in formulas_maestras.items():
                celda_f = get_or_create_cell(row_elem, indice_celdas, col, fila_num)
                nueva_formula = trasladar_formula(formula, 2, fila_num)
                escribir_formula(celda_f, nueva_formula)

        fila_final = 2 + len(datos) - 1
        for fila_num in range(fila_final + 1, limite_viejo + 1):
            limpiar_fila_sobrante(indice_filas, fila_num, max_col_fila)
        return fila_final

    else:  # INCONCERT MULTIVENTAS
        cantidad_filas_gestion = len(datos)
        fila_final = 2 + cantidad_filas_gestion - 1
        for fila_num in range(3, fila_final + 1):
            row_elem = get_row(indice_filas, fila_num)
            indice_celdas = indexar_celdas(row_elem)
            for col, formula in formulas_maestras.items():
                celda_f = get_or_create_cell(row_elem, indice_celdas, col, fila_num)
                nueva_formula = trasladar_formula(formula, 2, fila_num)
                escribir_formula(celda_f, nueva_formula)
        for fila_num in range(fila_final + 1, limite_viejo + 1):
            limpiar_fila_sobrante(indice_filas, fila_num, max_col_fila)
        return fila_final


def main() -> None:
    datos_brutas = leer_datos(BRUTAS)
    datos_gestion = leer_datos(GESTION)
    print(f"Filas nuevas -- brutas: {len(datos_brutas)} | gestionventas: {len(datos_gestion)}", flush=True)

    parser = etree.XMLParser(remove_blank_text=False)
    xml_modificado = {}

    with zipfile.ZipFile(ARCHIVO_CDM) as z:
        for nombre_hoja, ruta_interna in HOJAS.items():
            contenido = z.read(ruta_interna)
            root = etree.fromstring(contenido, parser=parser)
            datos = datos_gestion if nombre_hoja in ("Detalle", "INCONCERT MULTIVENTAS") else datos_brutas
            fila_final = procesar_hoja(root, nombre_hoja, datos, es_gestion=(nombre_hoja != "TABLA DE RESULTADOS INCONCERT"))
            xml_modificado[ruta_interna] = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
            print(f"{nombre_hoja}: procesada hasta la fila {fila_final}.", flush=True)

    temporal = ARCHIVO_CDM.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(ARCHIVO_CDM) as zin, zipfile.ZipFile(temporal, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename in xml_modificado:
                zout.writestr(item, xml_modificado[item.filename])
            else:
                zout.writestr(item, zin.read(item.filename))

    shutil.move(str(temporal), str(ARCHIVO_CDM))
    print("Guardado.", flush=True)


if __name__ == "__main__":
    main()
