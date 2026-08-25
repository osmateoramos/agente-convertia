"""
Limpia los reportes de Convertia ya descargados por descargar_reportes_convertia.py:
detecta automaticamente la fila real de encabezados (la guia de arriba puede variar de
largo) y normaliza las columnas de fecha (FCH_* / Fecha_*) quitandoles la hora. No
toca ninguna otra columna -- se mantienen todas, incluidas las NULL*, tal cual vienen
del export, porque estas tablas se pegan despues en otro archivo que espera esa
misma estructura de columnas. No depende de ningun Excel/Power Query ni de que la IA
revise archivo por archivo: es un pipeline determinista, el mismo archivo de entrada
siempre da el mismo resultado.

Requisitos (una sola vez):
    pip install openpyxl

Uso:
    python limpiar_reportes_convertia.py
        Limpia todos los .xlsx de esta carpeta que coincidan con los alias de
        reportes_config.json (BRUTAS_*, GESTIONVENTAS_*, etc.) y cuyo resultado
        procesado todavia no exista en Procesados/. Guarda cada uno en
        Procesados/<mismo nombre>.

    python limpiar_reportes_convertia.py --forzar
        Vuelve a limpiar aunque ya exista el archivo procesado correspondiente.

Una vez que un archivo crudo se limpia con exito (o si ya estaba limpio de antes),
el original se borra -- la idea es que en la carpeta principal solo queden los
scripts y la carpeta Procesados/, no los .xlsx crudos sueltos. Si un archivo da
ERROR al limpiarlo, el crudo NO se borra, para poder revisarlo a mano.
"""

import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_PATH = SCRIPT_DIR / "reportes_config.json"
CARPETA_PROCESADOS = SCRIPT_DIR / "Procesados"

MAX_FILAS_GUIA = 30  # si no aparece la fila separadora en blanco dentro de este margen, se considera formato inesperado
FORMATOS_FECHA_TEXTO = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y")


def cargar_alias_validos() -> set[str]:
    with open(CONFIG_PATH, encoding="utf-8") as f:
        config = json.load(f)
    alias = set()
    for datos_cuenta in config["cuentas"].values():
        for rep in datos_cuenta["reportes"]:
            alias.add(rep["alias"].upper())
    return alias


def es_archivo_de_reporte(nombre: str, alias_validos: set[str]) -> bool:
    if not nombre.lower().endswith(".xlsx"):
        return False
    prefijo = nombre.split("_", 1)[0]
    return prefijo in alias_validos


PATRON_NOMBRE = re.compile(r"^[A-Z0-9]+_(.+)_\d{4}-\d{2}-\d{2}_\d{4}-\d{2}-\d{2}\.xlsx$", re.IGNORECASE)


def extraer_cuenta_de_nombre(nombre: str) -> str | None:
    """Nombre esperado: {ALIAS}_{cuenta}_{fecha_inicio}_{fecha_fin}.xlsx
    (dos fechas ISO al final -- el rango consultado, no la fecha de descarga)."""
    m = PATRON_NOMBRE.match(nombre)
    return m.group(1) if m else None


def encontrar_fila_encabezado(filas):
    """filas es un iterador de tuplas (una fila del Excel cada una). Devuelve
    (numero_de_fila_del_encabezado, lista_de_encabezados, filas_de_guia_vistas).
    La fila de encabezado real es la primera fila no vacia inmediatamente despues
    de la primera fila completamente vacia -- asi el largo de la guia de arriba
    puede variar (7, 8, lo que sea) sin romper la deteccion."""
    vistas = []
    for i, fila in enumerate(filas, start=1):
        if i > MAX_FILAS_GUIA:
            raise ValueError(
                f"No aparecio una fila en blanco (separador antes del encabezado) "
                f"en las primeras {MAX_FILAS_GUIA} filas. Formato inesperado, revisar a mano."
            )
        vacia = all(v is None or str(v).strip() == "" for v in fila)
        if vacia and vistas:
            try:
                encabezado = next(filas)
            except StopIteration:
                raise ValueError("Se encontro la fila en blanco pero no hay fila de encabezado despues.")
            no_vacias = sum(1 for v in encabezado if v not in (None, ""))
            if no_vacias < 3:
                raise ValueError(f"La fila despues del separador tiene muy pocas columnas ({no_vacias}), no parece un encabezado real.")
            return i + 1, list(encabezado), vistas
        vistas.append(fila)
    raise ValueError("No aparecio ninguna fila en blanco separadora antes de agotar el margen de busqueda.")


def extraer_texto_guia(filas_guia, etiqueta: str) -> str | None:
    for fila in filas_guia:
        if not fila or fila[0] is None:
            continue
        texto = str(fila[0])
        if etiqueta in texto:
            return texto.split(etiqueta, 1)[1].strip()
    return None


def es_columna_fecha(nombre_columna) -> bool:
    if not nombre_columna:
        return False
    n = str(nombre_columna).upper()
    return n.startswith("FCH") or n.startswith("FECHA")


def limpiar_valor_fecha(valor):
    """None/vacio -> None. datetime/date -> date (sin hora). Texto parseable -> date.
    Cualquier otra cosa -> se devuelve tal cual (no se inventa un valor)."""
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    for fmt in FORMATOS_FECHA_TEXTO:
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return valor


def limpiar_archivo(ruta_entrada: Path, ruta_salida: Path) -> dict:
    wb_in = openpyxl.load_workbook(ruta_entrada, read_only=True, data_only=True)
    ws_in = wb_in[wb_in.sheetnames[0]]
    filas = ws_in.iter_rows(values_only=True)

    idx_encabezado, encabezados_originales, filas_guia = encontrar_fila_encabezado(filas)
    cuenta_guia = extraer_texto_guia(filas_guia, "Cuenta: ")
    resultados_totales_txt = extraer_texto_guia(filas_guia, "Resultados totales: ")

    columnas_fecha = {i for i, nom in enumerate(encabezados_originales) if es_columna_fecha(nom)}

    wb_out = openpyxl.Workbook(write_only=True)
    ws_out = wb_out.create_sheet("Datos")
    ws_out.append(encabezados_originales)

    total_filas = 0
    valores_fecha_no_parseables = 0
    for fila in filas:
        nueva_fila = []
        for i, valor in enumerate(fila):
            if i in columnas_fecha:
                limpio = limpiar_valor_fecha(valor)
                if valor not in (None, "") and not isinstance(limpio, date):
                    valores_fecha_no_parseables += 1
                nueva_fila.append(limpio)
            else:
                nueva_fila.append(valor)
        ws_out.append(nueva_fila)
        total_filas += 1

    wb_in.close()
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(ruta_salida)

    return {
        "fila_encabezado_detectada": idx_encabezado,
        "cuenta_en_guia": cuenta_guia,
        "resultados_totales_declarados": resultados_totales_txt,
        "filas_de_datos_escritas": total_filas,
        "columnas_fecha_limpiadas": len(columnas_fecha),
        "valores_fecha_no_parseables": valores_fecha_no_parseables,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--forzar",
        action="store_true",
        help="Volver a limpiar aunque ya exista el archivo procesado correspondiente en Procesados/",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    alias_validos = cargar_alias_validos()

    candidatos = sorted(
        p for p in SCRIPT_DIR.glob("*.xlsx") if es_archivo_de_reporte(p.name, alias_validos)
    )
    if not candidatos:
        print(f"No se encontro ningun .xlsx con alias conocido ({sorted(alias_validos)}) en {SCRIPT_DIR}.", flush=True)
        return

    for ruta_entrada in candidatos:
        ruta_salida = CARPETA_PROCESADOS / ruta_entrada.name
        if ruta_salida.exists() and not args.forzar:
            print(f"Ya existe Procesados/{ruta_entrada.name}, se omite la limpieza (usa --forzar para rehacerla).", flush=True)
            ruta_entrada.unlink()
            print(f"  Crudo borrado (ya no hace falta, el procesado ya existia): {ruta_entrada.name}", flush=True)
            continue

        print(f"--- Limpiando {ruta_entrada.name} ---", flush=True)
        try:
            info = limpiar_archivo(ruta_entrada, ruta_salida)
        except Exception as exc:
            print(f"ERROR limpiando {ruta_entrada.name}: {exc}", flush=True)
            print(f"  El crudo NO se borra, queda para revisarlo a mano: {ruta_entrada.name}", flush=True)
            continue

        cuenta_archivo = extraer_cuenta_de_nombre(ruta_entrada.name)
        if info["cuenta_en_guia"] and cuenta_archivo and info["cuenta_en_guia"] != cuenta_archivo:
            print(
                f"  AVISO: la cuenta en la guia del archivo ('{info['cuenta_en_guia']}') "
                f"no coincide con la del nombre de archivo ('{cuenta_archivo}').",
                flush=True,
            )

        totales = info["resultados_totales_declarados"]
        if totales and totales.isdigit() and int(totales) != info["filas_de_datos_escritas"]:
            print(
                f"  AVISO: el reporte declara 'Resultados totales: {totales}' pero se "
                f"escribieron {info['filas_de_datos_escritas']} filas de datos.",
                flush=True,
            )

        if info["valores_fecha_no_parseables"] > 0:
            print(f"  AVISO: {info['valores_fecha_no_parseables']} valores de columnas de fecha no se pudieron interpretar (se dejaron tal cual).", flush=True)

        print(
            f"  Encabezado real en fila {info['fila_encabezado_detectada']} | "
            f"{info['filas_de_datos_escritas']} filas | "
            f"{info['columnas_fecha_limpiadas']} columnas de fecha limpiadas | "
            f"todas las demas columnas se dejaron tal cual",
            flush=True,
        )
        print(f"  Guardado: Procesados/{ruta_entrada.name}", flush=True)

        ruta_entrada.unlink()
        print(f"  Crudo borrado: {ruta_entrada.name}", flush=True)


if __name__ == "__main__":
    main()
