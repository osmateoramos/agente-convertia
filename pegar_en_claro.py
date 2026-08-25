"""
PRUEBA: pega el brutas de Claro_Cr (01/07/2026 - 23/08/2026) en la hoja
'Brutas' de '08 Claro CR 23 de Agosto 2026.xlsx', reemplazando el historico
acumulado, desplegando la formula de la columna CA ('Mes Gestion') hasta
donde llegue la data nueva, actualizando INF!E18 a la fecha de corte nueva
(23/08/2026, mismo mes que ya tenia seleccionado el segmento 'Mes
instalacion': agosto) y marcando el cache de tablas dinamicas que se
alimenta de esta hoja para que se refresque solo al abrir el archivo en
Excel (no se puede recalcular un pivot sin Excel).

Mapeo de columnas confirmado a mano (25/08/2026): A:AU (47) y AZ:BY (26)
coinciden con Procesados/BRUTAS_Claro_Cr; las columnas AV:AY (4: DURATION_TIME,
ESTATUS, ULT_RESULTADO_MARCADORA, ULT_MARCADORA) y BZ (ULT_CAMPAÑA_INCONCERT)
no existen en el reporte actual -- quedan en blanco, misma politica que en
la prueba de WOW PERU.

La hoja 'BD' (dataset de ventas/instalaciones, fuente de otro pivot cache y
del slicer) es de otro sistema, no se toca.

Uso:
    python pegar_en_claro.py
"""

import shutil
import zipfile
from datetime import date
from pathlib import Path

import openpyxl
from lxml import etree
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.datetime import to_excel

CARPETA = Path(__file__).resolve().parent
ARCHIVO = CARPETA / "08 Claro CR 23 de Agosto 2026.xlsx"
BRUTAS = CARPETA / "Procesados" / "BRUTAS_Claro_Cr_2026-08-25.xlsx"

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

SHEET_BRUTAS = "xl/worksheets/sheet6.xml"
SHEET_INF = "xl/worksheets/sheet1.xml"
PIVOT_CACHE_BRUTAS = "xl/pivotCache/pivotCacheDefinition2.xml"  # sourceName Tabla3, ligado a 'Brutas'

COLUMNAS_QUE_COINCIDEN = set(range(1, 48)) | set(range(52, 78))  # A:AU y AZ:BY
COLUMNAS_TOTALES_HOJA = 78  # A:BZ
LIMITE_VIEJO = 7402
NUEVA_FECHA_CORTE = date(2026, 8, 23)


def q(tag: str) -> str:
    return f"{{{MAIN_NS}}}{tag}"


def leer_datos(ruta: Path) -> list[tuple]:
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return filas


def indexar_filas(sheet_data) -> dict[int, object]:
    return {int(row.get("r")): row for row in sheet_data if row.get("r") is not None}


def indexar_celdas(row_elem) -> dict[int, object]:
    indice = {}
    for c in row_elem:
        r_attr = c.get("r")
        if r_attr:
            indice[column_index_from_string("".join(ch for ch in r_attr if ch.isalpha()))] = c
    return indice


def get_or_create_row(sheet_data, indice_filas: dict, fila: int):
    row_elem = indice_filas.get(fila)
    if row_elem is not None:
        return row_elem
    nueva = etree.Element(q("row"))
    nueva.set("r", str(fila))
    insertado = False
    for hijo in sheet_data:
        r_attr = hijo.get("r")
        if r_attr and int(r_attr) > fila:
            hijo.addprevious(nueva)
            insertado = True
            break
    if not insertado:
        sheet_data.append(nueva)
    indice_filas[fila] = nueva
    return nueva


def get_or_create_cell(row_elem, indice_celdas: dict, col: int, fila: int, estilos_por_columna: dict | None = None):
    celda = indice_celdas.get(col)
    if celda is not None:
        return celda
    letra = get_column_letter(col)
    ref = f"{letra}{fila}"
    nueva = etree.Element(q("c"))
    nueva.set("r", ref)
    if estilos_por_columna and col in estilos_por_columna:
        nueva.set("s", estilos_por_columna[col])
    insertado = False
    for hijo in row_elem:
        r_attr = hijo.get("r")
        if r_attr:
            col_existente = column_index_from_string("".join(ch for ch in r_attr if ch.isalpha()))
            if col_existente > col:
                hijo.addprevious(nueva)
                insertado = True
                break
    if not insertado:
        row_elem.append(nueva)
    indice_celdas[col] = nueva
    return nueva


def limpiar_celda(celda) -> None:
    for hijo in list(celda):
        celda.remove(hijo)
    if "t" in celda.attrib:
        del celda.attrib["t"]


def escribir_valor(celda, valor) -> None:
    limpiar_celda(celda)
    if valor is None or valor == "":
        return
    if hasattr(valor, "hour"):  # datetime
        v = etree.SubElement(celda, q("v"))
        v.text = repr(to_excel(valor))
    elif hasattr(valor, "isoformat"):  # date
        v = etree.SubElement(celda, q("v"))
        v.text = repr(to_excel(valor))
    elif isinstance(valor, (int, float)):
        v = etree.SubElement(celda, q("v"))
        v.text = repr(valor)
    else:
        celda.set("t", "inlineStr")
        is_el = etree.SubElement(celda, q("is"))
        t_el = etree.SubElement(is_el, q("t"))
        texto = str(valor)
        if texto != texto.strip():
            t_el.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        t_el.text = texto


def escribir_formula(celda, formula_con_igual: str) -> None:
    limpiar_celda(celda)
    f_el = etree.SubElement(celda, q("f"))
    f_el.text = formula_con_igual[1:] if formula_con_igual.startswith("=") else formula_con_igual


def trasladar_formula(formula: str, fila_origen: int, fila_destino: int) -> str:
    return Translator(f"={formula}" if not formula.startswith("=") else formula,
                       origin=f"A{fila_origen}").translate_formula(f"A{fila_destino}")


def main() -> None:
    datos = leer_datos(BRUTAS)
    print(f"Filas nuevas de brutas: {len(datos)}", flush=True)

    parser = etree.XMLParser(remove_blank_text=False)

    with zipfile.ZipFile(ARCHIVO) as z:
        xml_brutas = z.read(SHEET_BRUTAS)
        xml_inf = z.read(SHEET_INF)
        xml_pivotcache = z.read(PIVOT_CACHE_BRUTAS)

    # ---------- Brutas ----------
    root_brutas = etree.fromstring(xml_brutas, parser=parser)
    sheet_data = root_brutas.find(q("sheetData"))
    indice_filas = indexar_filas(sheet_data)

    fila2 = get_or_create_row(sheet_data, indice_filas, 2)
    indice_celdas_fila2 = indexar_celdas(fila2)
    formula_ca_maestra = None
    celda_ca2 = indice_celdas_fila2.get(79)
    if celda_ca2 is not None:
        f_el = celda_ca2.find(q("f"))
        if f_el is not None:
            formula_ca_maestra = f_el.text

    # estilo (formato de numero/fecha) de cada columna, tomado de la fila 2 que ya
    # existia -- las filas que hay que crear de cero (mas alla de la fila 7402
    # vieja) necesitan este estilo a mano, si no las fechas salen como numero crudo
    estilos_por_columna = {col: c.get("s") for col, c in indice_celdas_fila2.items() if c.get("s")}

    registro0 = datos[0]
    for col in range(1, COLUMNAS_TOTALES_HOJA + 1):
        celda = get_or_create_cell(fila2, indice_celdas_fila2, col, 2, estilos_por_columna)
        if col in COLUMNAS_QUE_COINCIDEN:
            escribir_valor(celda, registro0[col - 1] if col - 1 < len(registro0) else None)
        else:
            limpiar_celda(celda)

    for i in range(1, len(datos)):
        fila_num = 2 + i
        row_elem = get_or_create_row(sheet_data, indice_filas, fila_num)
        indice_celdas = indexar_celdas(row_elem)
        registro = datos[i]
        for col in range(1, COLUMNAS_TOTALES_HOJA + 1):
            celda = get_or_create_cell(row_elem, indice_celdas, col, fila_num, estilos_por_columna)
            if col in COLUMNAS_QUE_COINCIDEN:
                escribir_valor(celda, registro[col - 1] if col - 1 < len(registro) else None)
            else:
                limpiar_celda(celda)
        if formula_ca_maestra:
            celda_ca = get_or_create_cell(row_elem, indice_celdas, 79, fila_num, estilos_por_columna)
            nueva_formula = trasladar_formula(formula_ca_maestra, 2, fila_num)
            escribir_formula(celda_ca, nueva_formula)

    fila_final = 2 + len(datos) - 1
    # limpiar filas sobrantes (por si el rango viejo llegara a ser mayor)
    for fila_num in range(fila_final + 1, LIMITE_VIEJO + 1):
        row_elem = indice_filas.get(fila_num)
        if row_elem is None:
            continue
        for c in list(row_elem):
            limpiar_celda(c)

    print(f"Brutas: {len(datos)} filas pegadas (fila 2 a {fila_final}), formula CA desplegada.", flush=True)

    xml_brutas_nuevo = etree.tostring(root_brutas, xml_declaration=True, encoding="UTF-8", standalone=True)

    # ---------- INF: actualizar E18 a la nueva fecha de corte ----------
    root_inf = etree.fromstring(xml_inf, parser=parser)
    sheet_data_inf = root_inf.find(q("sheetData"))
    indice_filas_inf = indexar_filas(sheet_data_inf)
    fila18 = get_or_create_row(sheet_data_inf, indice_filas_inf, 18)
    indice_celdas_18 = indexar_celdas(fila18)
    celda_e18 = get_or_create_cell(fila18, indice_celdas_18, 5, 18)
    escribir_valor(celda_e18, NUEVA_FECHA_CORTE)
    print("INF!E18 actualizado a 2026-08-23.", flush=True)
    xml_inf_nuevo = etree.tostring(root_inf, xml_declaration=True, encoding="UTF-8", standalone=True)

    # ---------- pivotCacheDefinition2 (Tabla3, alimentada por Brutas): refreshOnLoad ----------
    root_pc = etree.fromstring(xml_pivotcache, parser=parser)
    root_pc.set("refreshOnLoad", "1")
    print("pivotCacheDefinition2 marcado con refreshOnLoad=1 (Excel va a refrescar los pivots de 'Brutas' solo al abrir).", flush=True)
    xml_pivotcache_nuevo = etree.tostring(root_pc, xml_declaration=True, encoding="UTF-8", standalone=True)

    modificados = {
        SHEET_BRUTAS: xml_brutas_nuevo,
        SHEET_INF: xml_inf_nuevo,
        PIVOT_CACHE_BRUTAS: xml_pivotcache_nuevo,
    }

    temporal = ARCHIVO.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(ARCHIVO) as zin, zipfile.ZipFile(temporal, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename in modificados:
                zout.writestr(item, modificados[item.filename])
            else:
                zout.writestr(item, zin.read(item.filename))

    shutil.move(str(temporal), str(ARCHIVO))
    print("Guardado.", flush=True)


if __name__ == "__main__":
    main()
